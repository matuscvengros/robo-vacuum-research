#!/usr/bin/env python3
"""Preprocessing script: Extract text from HTML chat files and Excel spreadsheet."""

import os
from bs4 import BeautifulSoup
import openpyxl

BASE = "/Users/mako/Documents/RoboVacuumResearch"
OUT = os.path.join(BASE, "output")


def extract_html_chat(html_path, output_path, chat_name):
    """Extract readable text from a saved HTML chat page."""
    print(f"Processing {chat_name} HTML ({os.path.getsize(html_path):,} bytes)...")
    with open(html_path, "r", encoding="utf-8", errors="replace") as f:
        soup = BeautifulSoup(f.read(), "html.parser")

    # Remove script, style, and other non-content tags
    for tag in soup(["script", "style", "link", "meta", "noscript", "iframe", "svg", "path"]):
        tag.decompose()

    # Get text with reasonable whitespace handling
    text = soup.get_text(separator="\n", strip=False)

    # Clean up excessive blank lines
    lines = text.split("\n")
    cleaned = []
    blank_count = 0
    for line in lines:
        stripped = line.strip()
        if not stripped:
            blank_count += 1
            if blank_count <= 2:
                cleaned.append("")
        else:
            blank_count = 0
            cleaned.append(stripped)

    result = "\n".join(cleaned).strip()
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(result)
    print(f"  -> {output_path} ({len(result):,} chars, {len(cleaned):,} lines)")


def extract_excel(xlsx_path, output_path):
    """Extract Excel spreadsheet contents to Markdown."""
    print(f"Processing Excel ({os.path.getsize(xlsx_path):,} bytes)...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    md_lines = ["# Robot Vacuum Final Comparison - Excel Extract\n"]

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        md_lines.append(f"## Sheet: {sheet_name}\n")

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            md_lines.append("*(empty sheet)*\n")
            continue

        # Find max columns with data
        max_cols = 0
        for row in rows:
            for i in range(len(row) - 1, -1, -1):
                if row[i] is not None:
                    max_cols = max(max_cols, i + 1)
                    break

        if max_cols == 0:
            md_lines.append("*(empty sheet)*\n")
            continue

        # Build markdown table
        header = rows[0][:max_cols]
        header_strs = [str(c) if c is not None else "" for c in header]
        md_lines.append("| " + " | ".join(header_strs) + " |")
        md_lines.append("| " + " | ".join(["---"] * max_cols) + " |")

        for row in rows[1:]:
            cells = row[:max_cols]
            cell_strs = [str(c) if c is not None else "" for c in cells]
            md_lines.append("| " + " | ".join(cell_strs) + " |")

        md_lines.append("")

    result = "\n".join(md_lines)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(result)
    print(f"  -> {output_path} ({len(result):,} chars)")


if __name__ == "__main__":
    os.makedirs(OUT, exist_ok=True)

    # Extract Claude chat HTML
    extract_html_chat(
        os.path.join(BASE, "resources/Claude/Claude Chat.html"),
        os.path.join(OUT, "00_claude_chat_extract.txt"),
        "Claude"
    )

    # Extract Gemini chat HTML
    extract_html_chat(
        os.path.join(BASE, "resources/Gemini/Google Gemini.html"),
        os.path.join(OUT, "00_gemini_chat_extract.txt"),
        "Gemini"
    )

    # Extract Excel
    extract_excel(
        os.path.join(BASE, "resources/Claude/Robot_Vacuum_Final_Comparison_Mar2026.xlsx"),
        os.path.join(OUT, "00_excel_extract.md")
    )

    print("\nPreprocessing complete!")
